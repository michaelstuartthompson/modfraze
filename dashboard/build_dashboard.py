from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter

NAVY, CRIM, WINE, PLUM = "00224D", "FF204E", "A0153E", "5D0E41"
HDR_FILL = PatternFill("solid", start_color=NAVY)
HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BASE = Font(name="Arial", size=10)
INPUT = Font(name="Arial", size=10, color="0000FF")
thin = Side(style="thin", color="CCCCCC")
BORD = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

def make_sheet(name, headers, widths, color=None, freeze="A2"):
    ws = wb.create_sheet(name)
    if color: ws.sheet_properties.tabColor = color
    for i, h in enumerate(headers, 1):
        c = ws.cell(1, i, h); c.fill = HDR_FILL; c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = widths[i-1]
    ws.freeze_panes = freeze
    return ws

def dv(ws, col, options, rows=200):
    d = DataValidation(type="list", formula1='"' + ",".join(options) + '"', allow_blank=True)
    ws.add_data_validation(d); d.add(f"{col}2:{col}{rows}")

def style_rows(ws, ncols, rows=60):
    for r in range(2, rows + 2):
        for c in range(1, ncols + 1):
            cell = ws.cell(r, c)
            if cell.font.name != "Arial": cell.font = BASE

# ---------- Designs ----------
ws = make_sheet("Designs", ["design_id","title","collection","status","created_date","products_count","lifetime_revenue","notes"],
                [30,18,18,12,12,14,15,30], WINE)
ws["A2"]="MF-DES-20260701-NeonSprawl-V01"; ws["B2"]="Neon Sprawl"; ws["C2"]="MFC-Maximal-Vol1"
ws["D2"]="approved"; ws["E2"]="2026-07-01"; ws["H2"]="EXAMPLE ROW - replace/keep"
for r in range(2, 62):
    ws.cell(r,6).value = f'=IF($A{r}="","",COUNTIF(Products!$B:$B,$A{r}))'
    ws.cell(r,7).value = f'=IF($A{r}="","",SUMIF(Products!$B:$B,$A{r},Products!$K:$K))'
    ws.cell(r,7).number_format = '$#,##0.00'
dv(ws,"D",["draft","review","approved","published","archived"])
style_rows(ws,8)

# ---------- Products ----------
ws = make_sheet("Products", ["product_id","design_id","type","vendor","price","cost","margin","price_floor","platforms","listing_status","lifetime_revenue","lifetime_units"],
                [24,30,10,10,10,10,10,14,20,14,15,13], WINE)
ws["A2"]="MF-NEONSPRAWL-TEE"; ws["B2"]="MF-DES-20260701-NeonSprawl-V01"; ws["C2"]="TEE"; ws["D2"]="Printify"
ws["E2"]=29.99; ws["F2"]=11.50; ws["I2"]="shopify, tiktok_shop"; ws["J2"]="published"
ws["E2"].font = INPUT; ws["F2"].font = INPUT
for r in range(2, 62):
    ws.cell(r,7).value = f'=IF($E{r}="","",IF($E{r}=0,0,($E{r}-$F{r})/$E{r}))'
    ws.cell(r,7).number_format = '0.0%'
    ws.cell(r,8).value = f'=IF($E{r}="","",IF($E{r}<$F{r}*2.5,"BELOW FLOOR","OK"))'
    ws.cell(r,11).value = f'=IF($A{r}="","",SUMIF(Revenue!$D:$D,$A{r},Revenue!$F:$F))'
    ws.cell(r,11).number_format = '$#,##0.00'
    ws.cell(r,12).value = f'=IF($A{r}="","",SUMIF(Revenue!$D:$D,$A{r},Revenue!$E:$E))'
    for col in (5,6): ws.cell(r,col).number_format = '$#,##0.00'
dv(ws,"J",["draft","published","archived"])
dv(ws,"C",["TEE","HOOD","POSTER","MUG","TOTE","PHONE","STICKER","CANVAS"])
ws.conditional_formatting.add("H2:H61",
    CellIsRule(operator="equal", formula=['"BELOW FLOOR"'],
               fill=PatternFill("solid", start_color="F4CCCC"), font=Font(color="CC0000", bold=True)))
style_rows(ws,12)

# ---------- Shopify Listings ----------
ws = make_sheet("Shopify Listings", ["product_id","url","publish_date","status","views","add_to_cart","conversions","atc_rate","conv_rate"],
                [24,34,12,10,10,12,12,10,10], PLUM)
ws["A2"]="MF-NEONSPRAWL-TEE"; ws["B2"]="https://modfraze.com/products/neon-sprawl-tee"; ws["C2"]="2026-07-08"; ws["D2"]="live"
for r in range(2, 62):
    ws.cell(r,8).value = f'=IF($E{r}="","",IF($E{r}=0,0,$F{r}/$E{r}))'
    ws.cell(r,9).value = f'=IF($E{r}="","",IF($E{r}=0,0,$G{r}/$E{r}))'
    ws.cell(r,8).number_format = '0.0%'; ws.cell(r,9).number_format = '0.0%'
dv(ws,"D",["live","paused"]); style_rows(ws,9)

# ---------- TikTok Shop Listings ----------
ws = make_sheet("TikTok Shop Listings", ["product_id","url","publish_date","status","linked_videos","video_views","clicks","sales","click_rate"],
                [24,34,12,10,13,12,10,10,10], PLUM)
for r in range(2, 62):
    ws.cell(r,9).value = f'=IF($F{r}="","",IF($F{r}=0,0,$G{r}/$F{r}))'
    ws.cell(r,9).number_format = '0.00%'
dv(ws,"D",["live","paused"]); style_rows(ws,9)

# ---------- Queues ----------
ws = make_sheet("Etsy Queue", ["design_id","priority","why","earliest_date","status"], [30,10,34,13,10], "999999")
dv(ws,"B",["high","medium","low"]); dv(ws,"E",["queued","promoted","dropped"]); style_rows(ws,5)
ws = make_sheet("Pinterest Queue", ["product_id","pin_concept","earliest_date","status"], [24,34,13,10], "999999")
dv(ws,"D",["queued","promoted","dropped"]); style_rows(ws,4)

# ---------- Content Calendar ----------
ws = make_sheet("Content Calendar", ["post_id","platform","content_type","hook","product_id","scheduled_date","status","posted_url","views","clicks"],
                [28,11,18,32,24,13,10,30,10,10], CRIM)
ws["A2"]="MF-TT-20260710-POV-NeonSprawl"; ws["B2"]="tiktok"; ws["C2"]="product_reveal_video"
ws["D2"]="POV: your closet finally matches your personality"; ws["E2"]="MF-NEONSPRAWL-TEE"
ws["F2"]="2026-07-10"; ws["G2"]="planned"
dv(ws,"B",["tiktok","instagram","facebook","x","threads","reddit","linkedin","pinterest","email"])
dv(ws,"G",["planned","posted","skipped"])
ws.conditional_formatting.add("G2:G200", CellIsRule(operator="equal", formula=['"posted"'], fill=PatternFill("solid", start_color="D9EAD3")))
style_rows(ws,10)

# ---------- Ads ----------
ws = make_sheet("Ads", ["campaign_id","platform","objective","product_ids","daily_budget","total_cap","start_date","end_date","status"],
                [30,10,13,26,12,10,12,12,10], CRIM)
ws["A2"]="MF-AD-TT-202607-Conversions-01"; ws["B2"]="tiktok"; ws["C2"]="conversions"
ws["D2"]="MF-NEONSPRAWL-TEE"; ws["E2"]=10; ws["F2"]=70; ws["G2"]="2026-07-14"; ws["H2"]="2026-07-21"; ws["I2"]="active"
ws["E2"].font=INPUT; ws["F2"].font=INPUT
dv(ws,"B",["tiktok","meta","pinterest","other"]); dv(ws,"I",["draft","active","paused","ended"])
for r in range(2,62):
    for col in (5,6): ws.cell(r,col).number_format = '$#,##0'
style_rows(ws,9)

# ---------- Ad Performance ----------
ws = make_sheet("Ad Performance", ["ad_id","campaign_id","review_date","days_running","spend","impressions","clicks","CTR","CPC","conversions","revenue","ROAS","DECISION","next_action"],
                [32,30,12,12,10,12,10,9,9,12,11,8,11,36], CRIM)
ws["A2"]="MF-AD-TT-202607-Conversions-01-A"; ws["B2"]="MF-AD-TT-202607-Conversions-01"
ws["C2"]="2026-07-17"; ws["D2"]=3; ws["E2"]=30.0; ws["F2"]=14200; ws["G2"]=168; ws["J2"]=2; ws["K2"]=59.98
ws["N2"]="Test new hook on same creative; keep budget flat 3 more days"
for r in range(2, 122):
    ws.cell(r,8).value  = f'=IF($F{r}="","",IF($F{r}=0,0,$G{r}/$F{r}))'
    ws.cell(r,9).value  = f'=IF($G{r}="","",IF($G{r}=0,0,$E{r}/$G{r}))'
    ws.cell(r,12).value = f'=IF($E{r}="","",IF($E{r}=0,0,$K{r}/$E{r}))'
    ws.cell(r,13).value = (f'=IF($A{r}="","",IF(AND($E{r}<15,$D{r}<3),"WAIT",'
                           f'IF(AND($L{r}>3,$H{r}>=0.01),"PUSH",'
                           f'IF($L{r}>=2,"ITERATE",IF($L{r}>=1.5,"PAUSE","KILL")))))')
    ws.cell(r,8).number_format='0.00%'; ws.cell(r,9).number_format='$#,##0.00'
    ws.cell(r,5).number_format='$#,##0.00'; ws.cell(r,11).number_format='$#,##0.00'
    ws.cell(r,12).number_format='0.00"x"'
    ws.cell(r,13).font = Font(name="Arial", size=10, bold=True)
    ws.cell(r,13).alignment = Alignment(horizontal="center")
rng = "M2:M121"
ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"PUSH"'],    fill=PatternFill("solid", start_color="B6D7A8")))
ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"ITERATE"'], fill=PatternFill("solid", start_color="FFE599")))
ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"PAUSE"'],   fill=PatternFill("solid", start_color="F9CB9C")))
ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"KILL"'],    fill=PatternFill("solid", start_color="EA9999")))
ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"WAIT"'],    fill=PatternFill("solid", start_color="D9D9D9")))
style_rows(ws,14,120)

# ---------- Email Campaigns ----------
ws = make_sheet("Email Campaigns", ["email_id","segment","subject_line","send_date","status","sent","opens","clicks","revenue","open_rate","click_rate"],
                [30,16,36,12,11,9,9,9,11,10,10], PLUM)
for r in range(2,62):
    ws.cell(r,10).value = f'=IF($F{r}="","",IF($F{r}=0,0,$G{r}/$F{r}))'
    ws.cell(r,11).value = f'=IF($F{r}="","",IF($F{r}=0,0,$H{r}/$F{r}))'
    ws.cell(r,10).number_format='0.0%'; ws.cell(r,11).number_format='0.0%'
    ws.cell(r,9).number_format='$#,##0.00'
dv(ws,"E",["draft","scheduled","sent"])
dv(ws,"B",["all_subscribers","buyers","non_buyers","vip_collectors","inactive_60d"])
style_rows(ws,11)

# ---------- Audience Segments ----------
ws = make_sheet("Audience Segments", ["segment","size","size_prev_week","growth_7d","last_campaign"], [18,10,15,11,28], PLUM)
for seg in ["all_subscribers","buyers","non_buyers","vip_collectors"]:
    ws.append([seg,"","","",""])
for r in range(2,42):
    ws.cell(r,4).value = f'=IF($B{r}="","",$B{r}-$C{r})'
style_rows(ws,5)

# ---------- Revenue ----------
ws = make_sheet("Revenue", ["date","order_id","platform","product_id","units","gross","fees","cost","net"],
                [12,16,12,24,8,11,10,10,11], "38761D")
ws["A2"]="2026-07-16"; ws["B2"]="EX-1001"; ws["C2"]="shopify"; ws["D2"]="MF-NEONSPRAWL-TEE"
ws["E2"]=1; ws["F2"]=29.99; ws["G2"]=1.17; ws["H2"]=11.50
for r in range(2,202):
    ws.cell(r,9).value = f'=IF($F{r}="","",$F{r}-$G{r}-$H{r})'
    for col in (6,7,8,9): ws.cell(r,col).number_format='$#,##0.00'
dv(ws,"C",["shopify","tiktok_shop","etsy","other"])
style_rows(ws,9,200)

# ---------- Costs ----------
ws = make_sheet("Costs", ["date","category","amount","note"], [12,12,11,36], "38761D")
for r in range(2,202): ws.cell(r,3).number_format='$#,##0.00'
dv(ws,"B",["tools","ads","fees","samples","other"]); style_rows(ws,4,200)

# ---------- Experiments ----------
ws = make_sheet("Experiments", ["exp_id","hypothesis","metric","start","end","status","result","learning"],
                [14,40,16,12,12,10,20,40], "999999")
dv(ws,"F",["running","done"]); style_rows(ws,8)

# ---------- Automation Log ----------
ws = make_sheet("Automation Log", ["date","automation","trigger","result","manual_fallback_used","note"],
                [12,28,24,10,20,32], "999999")
dv(ws,"D",["ok","failed"]); style_rows(ws,6)

# ---------- Portfolio Evidence Log ----------
ws = make_sheet("Portfolio Evidence Log", ["date","artifact","skill_demonstrated","file_path_or_screenshot","note"],
                [12,30,26,38,28], CRIM)
ws["A2"]="2026-07-01"; ws["B2"]="ModFraze_Dashboard v1 built"
ws["C2"]="dashboard design, data architecture"; ws["D2"]="08_Analytics_Dashboard/Weekly_Snapshots/Week01"
style_rows(ws,5,120)

# ---------- Archive ----------
ws = make_sheet("Archive", ["original_tab","record_id","archive_date","reason"], [18,32,13,40], "666666")
style_rows(ws,4)

# ---------- Dashboard (built last, referencing all tabs; placed first) ----------
dash = wb["Sheet"]; dash.title = "Dashboard"; dash.sheet_properties.tabColor = CRIM
wb.move_sheet("Dashboard", offset=-(len(wb.sheetnames)-1))
d = dash
d.column_dimensions["A"].width = 2
d.column_dimensions["B"].width = 34
d.column_dimensions["C"].width = 16
d.column_dimensions["D"].width = 44
d["B2"]="MODFRAZE COMMAND DASHBOARD"; d["B2"].font=Font(name="Arial",bold=True,size=16,color=NAVY)
d["B3"]='=TEXT(TODAY(),"dddd, mmmm d, yyyy")'; d["B3"].font=Font(name="Arial",size=11,color="666666")
d["D2"]="Daily habit: 5 minutes here, then execute."; d["D2"].font=Font(name="Arial",italic=True,size=10,color=CRIM)

def section(row, label):
    c=d.cell(row,2,label); c.fill=HDR_FILL; c.font=HDR_FONT
    d.cell(row,3,"").fill=HDR_FILL; d.cell(row,4,"SOURCE / NOTE").fill=HDR_FILL
    d.cell(row,4).font=Font(name="Arial",bold=True,italic=True,color="FFFFFF",size=9)

def kpi(row, label, formula, fmt=None, note="", manual=False):
    d.cell(row,2,label).font=Font(name="Arial",size=10)
    c=d.cell(row,3,formula)
    c.font = INPUT if manual else Font(name="Arial",size=10,bold=True)
    if fmt: c.number_format=fmt
    c.alignment=Alignment(horizontal="right")
    d.cell(row,4,note).font=Font(name="Arial",size=9,color="888888")

section(5,"REVENUE")
kpi(6,"Revenue (7d)",  "=SUMIFS(Revenue!F:F,Revenue!A:A,\">=\"&TEXT(TODAY()-7,\"yyyy-mm-dd\"))", '$#,##0.00', "Revenue tab · gross")
kpi(7,"Revenue (30d)", "=SUMIFS(Revenue!F:F,Revenue!A:A,\">=\"&TEXT(TODAY()-30,\"yyyy-mm-dd\"))", '$#,##0.00', "Revenue tab · gross")
kpi(8,"Net revenue (30d)", "=SUMIFS(Revenue!I:I,Revenue!A:A,\">=\"&TEXT(TODAY()-30,\"yyyy-mm-dd\"))", '$#,##0.00', "gross − fees − cost")
kpi(9,"Orders (7d)",   "=COUNTIFS(Revenue!A:A,\">=\"&TEXT(TODAY()-7,\"yyyy-mm-dd\"))", '#,##0', "Revenue tab")
kpi(10,"Gross margin % (30d)", "=IF(C7=0,0,C8/C7)", '0.0%', "net / gross")
kpi(11,"Shopify revenue (30d)", "=SUMIFS(Revenue!F:F,Revenue!A:A,\">=\"&TEXT(TODAY()-30,\"yyyy-mm-dd\"),Revenue!C:C,\"shopify\")", '$#,##0.00', "")
kpi(12,"TikTok Shop revenue (30d)", "=SUMIFS(Revenue!F:F,Revenue!A:A,\">=\"&TEXT(TODAY()-30,\"yyyy-mm-dd\"),Revenue!C:C,\"tiktok_shop\")", '$#,##0.00', "")
kpi(13,"Best platform (30d)", "=IF(AND(C11=0,C12=0),\"—\",IF(C11>=C12,\"Shopify\",\"TikTok Shop\"))", None, "")

section(15,"CATALOG")
kpi(16,"Designs approved/published", "=COUNTIF(Designs!D:D,\"approved\")+COUNTIF(Designs!D:D,\"published\")", '#,##0', "Designs tab")
kpi(17,"Products live", "=COUNTIF(Products!J:J,\"published\")", '#,##0', "Products tab")
kpi(18,"Top product (lifetime rev)", "=IFERROR(IF(MAX(Products!K2:K61)=0,\"—\",INDEX(Products!A2:A61,MATCH(MAX(Products!K2:K61),Products!K2:K61,0))),\"—\")", None, "Products tab")
kpi(19,"Top design (lifetime rev)", "=IFERROR(IF(MAX(Designs!G2:G61)=0,\"—\",INDEX(Designs!B2:B61,MATCH(MAX(Designs!G2:G61),Designs!G2:G61,0))),\"—\")", None, "Designs tab")
kpi(20,"⚠ Products below 2.5x price floor", "=COUNTIF(Products!H:H,\"BELOW FLOOR\")", '#,##0', "must be 0")

section(22,"ADS")
kpi(23,"Ad spend (7d)", "=SUMIFS('Ad Performance'!E:E,'Ad Performance'!C:C,\">=\"&TEXT(TODAY()-7,\"yyyy-mm-dd\"))", '$#,##0.00', "Ad Performance tab")
kpi(24,"Blended ROAS (7d)", "=IF(C23=0,0,SUMIFS('Ad Performance'!K:K,'Ad Performance'!C:C,\">=\"&TEXT(TODAY()-7,\"yyyy-mm-dd\"))/C23)", '0.00"x"', "ad revenue / spend")
kpi(25,"Active campaigns", "=COUNTIF(Ads!I:I,\"active\")", '#,##0', "max 3 during MVP")
kpi(26,"🟢 PUSH decisions", "=COUNTIF('Ad Performance'!M:M,\"PUSH\")", '#,##0', "scale these")
kpi(27,"🔴 KILL decisions", "=COUNTIF('Ad Performance'!M:M,\"KILL\")", '#,##0', "archive today")

section(29,"AUDIENCE & CONTENT")
kpi(30,"Email subscribers", "", '#,##0', "MANUAL: from Mailchimp, weekly", manual=True)
kpi(31,"Subscriber growth (7d)", "=IFERROR(INDEX('Audience Segments'!D:D,MATCH(\"all_subscribers\",'Audience Segments'!A:A,0)),0)", '#,##0', "Audience Segments tab")
kpi(32,"Posts published (7d)", "=COUNTIFS('Content Calendar'!G:G,\"posted\",'Content Calendar'!F:F,\">=\"&TEXT(TODAY()-7,\"yyyy-mm-dd\"))", '#,##0', "target: 3+ TikToks")
kpi(33,"Posts planned, next 7d", "=COUNTIFS('Content Calendar'!G:G,\"planned\",'Content Calendar'!F:F,\"<=\"&TEXT(TODAY()+7,\"yyyy-mm-dd\"))", '#,##0', "Content Calendar")

section(35,"SHOPIFY WEEKLY INPUTS (manual, Mondays)")
kpi(36,"Sessions (7d)", "", '#,##0', "MANUAL: Shopify analytics", manual=True)
kpi(37,"Add-to-carts (7d)", "", '#,##0', "MANUAL", manual=True)
kpi(38,"Reached checkout (7d)", "", '#,##0', "MANUAL", manual=True)
kpi(39,"Add-to-cart rate", "=IF(OR(C36=\"\",C36=0),0,C37/C36)", '0.0%', "")
kpi(40,"Conversion rate", "=IF(OR(C36=\"\",C36=0),0,C9/C36)", '0.0%', "orders(7d) / sessions")
kpi(41,"Abandoned checkout rate", "=IF(OR(C38=\"\",C38=0),0,MAX(0,(C38-C9)/C38))", '0.0%', "")
kpi(42,"Email signup rate", "=IF(OR(C36=\"\",C36=0),0,C31/C36)", '0.0%', "new subs / sessions")

section(44,"DAILY CHECK (2 minutes)")
for i, item in enumerate(["1. Any KILL or PUSH rows above? Act on them.",
                          "2. Is today's content posted or scheduled?",
                          "3. One Portfolio Evidence Log entry made?",
                          "4. Anything below price floor? Fix now.",
                          "5. Close the sheet. Go make things."], start=45):
    d.cell(i,2,item).font=Font(name="Arial",size=10,color="444444")

for row in d.iter_rows(min_row=5,max_row=42,min_col=2,max_col=4):
    for cell in row: cell.border=BORD

wb.save("/home/claude/modfraze/ModFraze_Dashboard.xlsx")
print("saved", wb.sheetnames)
